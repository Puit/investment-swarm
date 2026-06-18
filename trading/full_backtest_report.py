"""
Genera un reporte XLSX completo con backtests por categoría de stock.
Hojas: MAX (2020→2026) + una por cada año (2020-2026).
"""
from __future__ import annotations
import io
from datetime import date
from typing import Callable, Optional

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Categorías de stocks ────────────────────────────────────────────────────
CATEGORIES: dict[str, dict] = {
    "Momentum/AI": {
        "tickers":     ["NVDA", "META", "AMD", "TSLA", "CRWD"],
        "description": "Máxima volatilidad, crecimiento explosivo",
    },
    "Trend/Mature-Tech": {
        "tickers":     ["MSFT", "GOOG", "AAPL", "ORCL", "ADBE"],
        "description": "Tendencia estable, drawdowns moderados",
    },
    "Healthcare": {
        "tickers":     ["LLY", "UNH", "ABBV"],
        "description": "Mezcla growth y defensivo",
    },
    "Cyclical": {
        "tickers":     ["CAT", "DE", "XOM", "CVX", "JPM", "GS", "FCX", "NUE"],
        "description": "Muy sensibles al ciclo económico",
    },
    "Payments": {
        "tickers":     ["V", "MA"],
        "description": "Tendencia estable con moat",
    },
    "Defensive": {
        "tickers":     ["JNJ", "PG", "KO", "WMT", "NEE"],
        "description": "Baja volatilidad, dividendos",
    },
}

# ── Periodos a testear ────────────────────────────────────────────────────────
MAX_START = date(2020, 1, 1)
MAX_END   = date(2026, 6, 18)

PERIODS: list[dict] = [
    {"name": "MAX",  "start": MAX_START, "end": MAX_END},
    {"name": "2020", "start": date(2020, 1, 1), "end": date(2020, 12, 31)},
    {"name": "2021", "start": date(2021, 1, 1), "end": date(2021, 12, 31)},
    {"name": "2022", "start": date(2022, 1, 1), "end": date(2022, 12, 31)},
    {"name": "2023", "start": date(2023, 1, 1), "end": date(2023, 12, 31)},
    {"name": "2024", "start": date(2024, 1, 1), "end": date(2024, 12, 31)},
    {"name": "2025", "start": date(2025, 1, 1), "end": date(2025, 12, 31)},
    {"name": "2026", "start": date(2026, 1, 1), "end": MAX_END},
]

# ── Runner ───────────────────────────────────────────────────────────────────

def _run_one(
    category: str,
    period: dict,
    capital: float,
    cost_pct: float,
    cb: Optional[Callable[[str], None]] = None,
) -> dict:
    """Ejecuta un backtest para una categoría y un periodo, devuelve métricas."""
    from trading.backtest_runner import BacktestRunner

    tickers = CATEGORIES[category]["tickers"]
    pname   = period["name"]

    if cb:
        cb(f"{pname} · {category} ({', '.join(tickers)})…")

    r = BacktestRunner.run_backtest(
        date_range="Custom",
        initial_capital=capital,
        transaction_cost_pct=cost_pct,
        custom_start=period["start"],
        custom_end=period["end"],
        tickers=tickers,
        progress_callback=None,
    )

    if not r["success"]:
        return {"error": r.get("error", "unknown")}

    m   = r["metrics"]
    p   = m["performance"]
    t   = m["trading"]
    c   = m["comparison"]
    reg = m.get("regime", {})

    return {
        "capital_final": float(p["final_value"]),
        "return_pct":    float(p["total_return_pct"]),
        "max_dd":        float(p["max_drawdown_pct"]),
        "sharpe":        float(p["sharpe_ratio"]),
        "buys":          int(t["total_buys"]),
        "sells":         int(t["total_sells"]),
        "positions":     int(t.get("open_positions", 0)),
        "win_rate":      float(t["win_rate_pct"]),
        "stop_loss":     int(t["stop_loss_count"]),
        "bot_pct":       float(c["bot_return_pct"]),
        "bnh_pct":       float(c["buyhold_return_pct"]),
        "spy_pct":       float(c["spy_return_pct"]),
    }


def run_full_report(
    capital: float = 5_000.0,
    cost_pct: float = 0.001,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> dict[str, dict[str, dict]]:
    """
    Ejecuta todos los periodos × categorías.
    Devuelve: { "MAX": { "Momentum/AI": {...}, ...}, "2020": {...}, ... }
    """
    results: dict[str, dict[str, dict]] = {}
    total = len(PERIODS) * len(CATEGORIES)
    done  = 0

    for period in PERIODS:
        pname = period["name"]
        results[pname] = {}
        for cat in CATEGORIES:
            def _cb(msg, p=pname, c=cat, d=done, tot=total):
                pct = int(d / tot * 100)
                if progress_callback:
                    progress_callback(f"[{pct}%] {p} · {c}: {msg}")
            results[pname][cat] = _run_one(cat, period, capital, cost_pct, _cb)
            done += 1

    return results


# ── XLSX writer ──────────────────────────────────────────────────────────────

# Paleta de colores
_C_HEADER_TITLE = "1F3864"   # azul muy oscuro
_C_HEADER_SUB   = "2E75B6"   # azul medio
_C_COL_HEAD     = "D6E4F7"   # azul muy claro
_C_ALT_ROW      = "EEF4FB"   # fila alternada
_C_GREEN        = "E2EFDA"   # fondo si bot > bnh
_C_RED          = "FFDDD6"   # fondo si bot < 0
_C_POSITIVE     = "375623"   # texto retorno positivo
_C_NEGATIVE     = "9C0006"   # texto retorno negativo
_WHITE          = "FFFFFF"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=None, size=10, italic=False):
    kw: dict = {"bold": bold, "size": size, "italic": italic}
    if color:
        kw["color"] = color
    return Font(**kw)


def _border_thin() -> Border:
    s = Side(style="thin", color="B0C4DE")
    return Border(left=s, right=s, top=s, bottom=s)


def _pct_str(val: float) -> str:
    sign = "+" if val >= 0 else ""
    return f"{sign}{val:.2f}%"


def _write_sheet(
    ws,
    period: dict,
    period_results: dict[str, dict],
    capital: float,
    cost_pct: float,
):
    """Escribe una hoja con la tabla del periodo dado."""

    start_str = period["start"].strftime("%Y/%m/%d")
    end_str   = period["end"].strftime("%Y/%m/%d")
    cost_str  = f"{cost_pct * 100:.2f}%"
    cap_str   = f"{capital:,.2f}$"

    # ── Fila 1: título general ────────────────────────────────────────────
    header_text = (
        f"Pruebas realizadas desde {start_str} al {end_str}, "
        f"con un capital de {cap_str} y una comisión del {cost_str}"
    )
    ws.merge_cells("A1:O1")
    c = ws["A1"]
    c.value = header_text
    c.fill  = _fill(_C_HEADER_TITLE)
    c.font  = _font(bold=True, color=_WHITE, size=11)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # ── Fila 2: grupos de columnas ────────────────────────────────────────
    # Columnas A-C vacías (categoría/tickers/descripción)
    ws.merge_cells("D2:H2")   # Resumen de resultados  (D=capital, E=return, F=dd, G=sharpe)
    ws.merge_cells("I2:M2")   # Actividad del bot       (I=buys, J=sells, K=pos, L=wr, M=sl)
    ws.merge_cells("N2:O2")   # Comparación benchmarks  (N=bnh, O=spy)   + col B=Bot (col N)

    groups = {
        "D2": ("Resumen de resultados",     _C_HEADER_SUB),
        "I2": ("Actividad del bot",          _C_HEADER_SUB),
        "N2": ("Comparación con Benchmarks", _C_HEADER_SUB),
    }
    for cell_ref, (text, color) in groups.items():
        c = ws[cell_ref]
        c.value = text
        c.fill  = _fill(color)
        c.font  = _font(bold=True, color=_WHITE, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 18

    # ── Fila 3: cabeceras de columna ──────────────────────────────────────
    headers = [
        "Categoría", "Tickers", "Por qué incluirlos",
        "Capital final", "Retorno %", "Max Drawdown", "Sharpe Ratio",
        "Compra", "Venta", "Posiciones", "Win Rate", "Stop Loss",
        "Bot", "Buy&Hold", "SPY",
    ]
    for col_i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col_i)
        c.value = h
        c.fill  = _fill(_C_COL_HEAD)
        c.font  = _font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border_thin()
    ws.row_dimensions[3].height = 32

    # ── Filas de datos ────────────────────────────────────────────────────
    for row_i, (cat, meta) in enumerate(CATEGORIES.items(), start=4):
        m   = period_results.get(cat, {})
        alt = (row_i % 2 == 0)
        bg  = _C_ALT_ROW if alt else _WHITE

        tickers_str = ", ".join(meta["tickers"])
        desc        = meta["description"]

        if "error" in m:
            row_data = [cat, tickers_str, desc] + ["—"] * 12
        else:
            row_data = [
                cat,
                tickers_str,
                desc,
                f"${m['capital_final']:,.2f}",
                _pct_str(m["return_pct"]),
                f"{m['max_dd']:.2f}%",
                f"{m['sharpe']:.2f}",
                m["buys"],
                m["sells"],
                m["positions"],
                f"{m['win_rate']:.2f}%",
                m["stop_loss"],
                _pct_str(m["bot_pct"]),
                _pct_str(m["bnh_pct"]),
                _pct_str(m["spy_pct"]),
            ]

        for col_i, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.fill   = _fill(bg)
            c.font   = _font(size=9)
            c.border = _border_thin()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(col_i <= 3))

            # Color semántico en columnas de retorno
            if col_i == 5 and "error" not in m:  # Retorno %
                pct = m["return_pct"]
                c.font = _font(bold=True, color=_C_POSITIVE if pct >= 0 else _C_NEGATIVE, size=9)
            if col_i == 13 and "error" not in m:  # Bot
                pct = m["bot_pct"]
                c.font = _font(bold=True, color=_C_POSITIVE if pct >= 0 else _C_NEGATIVE, size=9)

        ws.row_dimensions[row_i].height = 20

    # Alineación izquierda para texto libre
    for row_i in range(4, 4 + len(CATEGORIES)):
        ws.cell(row=row_i, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_i, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row_i, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Anchos de columna ─────────────────────────────────────────────────
    col_widths = [22, 30, 40, 14, 11, 12, 10, 8, 8, 10, 10, 10, 11, 11, 11]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze filas de cabecera
    ws.freeze_panes = "A4"


def generate_xlsx(
    all_results: dict[str, dict[str, dict]],
    capital: float = 5_000.0,
    cost_pct: float = 0.001,
) -> bytes:
    """
    Genera el XLSX con una hoja por periodo.
    Devuelve los bytes del fichero para descarga.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quita la hoja por defecto

    for period in PERIODS:
        pname    = period["name"]
        ws       = wb.create_sheet(title=pname)
        p_result = all_results.get(pname, {})
        _write_sheet(ws, period, p_result, capital, cost_pct)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
